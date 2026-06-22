from aiogram import Router, F, Bot
from aiogram.filters import Command
from aiogram.types import Message, CallbackQuery
from aiogram.fsm.context import FSMContext

from app.config import SUPER_ADMIN_ID
from app.database.crud import (
    get_admin, get_all_admins, add_admin, remove_admin, update_admin_role,
    get_all_vacancies, get_vacancy, create_vacancy, toggle_vacancy, delete_vacancy,
    update_vacancy, get_applications, get_application, delete_application,
    get_all_subscribed_users,
    get_setting, set_setting,
)
from app.keyboards.inline import (
    admin_main_keyboard, admin_settings_keyboard,
    admin_vacancies_keyboard, admin_vacancy_detail_keyboard, vacancy_delete_confirm_keyboard,
    vacancy_edit_field_keyboard, vacancy_post_menu_keyboard, applications_post_menu_keyboard,
    admin_list_keyboard, admin_detail_keyboard, admin_roles_keyboard,
    admin_remove_confirm_keyboard, ROLE_LABELS,
)
from app.states.admin_state import (
    AddVacancyState, AddAdminState, EditAdminState, EditVacancyState,
    BotSettingsState, SearchApplicationState,
)

router = Router()


async def get_role(user_id: int) -> str | None:
    if user_id == SUPER_ADMIN_ID:
        return "super_admin"
    admin = await get_admin(user_id)
    return admin.role if admin else None


def is_hr(role: str) -> bool:
    return role in ("hr_admin", "super_admin")


# ── /admin ─────────────────────────────────────────────────────────────────

@router.message(Command("admin"))
async def admin_panel(message: Message):
    role = await get_role(message.from_user.id)
    if not role:
        await message.answer("❌ Sizda admin huquqi yo'q.")
        return
    await message.answer(
        f"👤 Admin paneli | Rol: <b>{role}</b>",
        parse_mode="HTML",
        reply_markup=admin_main_keyboard(role)
    )


@router.callback_query(lambda c: c.data == "admin:back")
async def admin_back(callback: CallbackQuery):
    role = await get_role(callback.from_user.id)
    if not role:
        await callback.answer("❌ Ruxsat yo'q.")
        return
    await callback.message.edit_text(
        f"👤 Admin paneli | Rol: <b>{role}</b>",
        parse_mode="HTML",
        reply_markup=admin_main_keyboard(role)
    )
    await callback.answer()


# ── Statistika ─────────────────────────────────────────────────────────────

@router.message(Command("stats"))
@router.callback_query(lambda c: c.data == "admin:stats")
async def show_stats(update, **kwargs):
    msg = update if isinstance(update, Message) else update.message
    uid = update.from_user.id
    role = await get_role(uid)
    if not role:
        await msg.answer("❌ Ruxsat yo'q.")
        return
    users     = await get_all_subscribed_users()
    vacancies = await get_all_vacancies()
    apps      = await get_applications()
    await msg.answer(
        f"📊 <b>Statistika</b>\n\n"
        f"👥 Foydalanuvchilar: {len(users)}\n"
        f"💼 Vakansiyalar: {len(vacancies)}\n"
        f"📁 Arizalar: {len(apps)}",
        parse_mode="HTML"
    )
    if isinstance(update, CallbackQuery):
        await update.answer()


# ── Vakansiyalar ───────────────────────────────────────────────────────────

@router.callback_query(lambda c: c.data == "admin:vacancies")
async def admin_vacancies(callback: CallbackQuery):
    role = await get_role(callback.from_user.id)
    if not is_hr(role):
        await callback.answer("❌ Ruxsat yo'q.")
        return
    vacancies = await get_all_vacancies()
    await callback.message.edit_text(
        "💼 Vakansiyalar:",
        reply_markup=admin_vacancies_keyboard(vacancies)
    )
    await callback.answer()


@router.callback_query(lambda c: c.data.startswith("admin_vacancy:"))
async def admin_vacancy_detail(callback: CallbackQuery, state: FSMContext):
    role = await get_role(callback.from_user.id)
    if not is_hr(role):
        await callback.answer("❌ Ruxsat yo'q.")
        return
    val = callback.data.split(":")[1]
    if val == "new":
        await state.set_state(AddVacancyState.title)
        await callback.message.answer("Yangi vakansiya nomi (lavozim):")
        await callback.answer()
        return
    v = await get_vacancy(int(val))
    text = (
        f"💼 <b>{v.title}</b>\n\n"
        f"📋 Talablar:\n{v.requirements or '—'}\n\n"
        f"💰 Ish haqi: {v.salary or 'Kelishiladi'}\n"
        f"Holat: {'🟢 Ochiq' if v.active else '🔴 Yopiq'}"
    )
    await callback.message.answer(
        text, parse_mode="HTML",
        reply_markup=admin_vacancy_detail_keyboard(v.id, v.active)
    )
    await callback.answer()


@router.message(AddVacancyState.title)
async def new_vacancy_title(message: Message, state: FSMContext):
    await state.update_data(title=message.text)
    await state.set_state(AddVacancyState.requirements)
    await message.answer("📋 Talablarni kiriting:")


@router.message(AddVacancyState.requirements)
async def new_vacancy_req(message: Message, state: FSMContext):
    await state.update_data(requirements=message.text)
    await state.set_state(AddVacancyState.salary)
    from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
    skip_kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="⏭ O'tkazib yuborish (Kelishiladi)", callback_data="vacancy_salary_skip")
    ]])
    await message.answer("💰 Ish haqi miqdorini kiriting:\n<i>Masalan: 3 000 000 so'm yoki O'tkazib yuboring</i>",
                         parse_mode="HTML", reply_markup=skip_kb)


@router.callback_query(AddVacancyState.salary, lambda c: c.data == "vacancy_salary_skip")
async def new_vacancy_salary_skip(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    v = await create_vacancy(data["title"], data["requirements"], salary=None)
    await state.clear()
    await callback.message.answer(f"✅ Vakansiya yaratildi: <b>{v.title}</b>", parse_mode="HTML")
    await callback.answer()


@router.message(AddVacancyState.salary)
async def new_vacancy_salary(message: Message, state: FSMContext):
    data = await state.get_data()
    v = await create_vacancy(data["title"], data["requirements"], salary=message.text.strip())
    await state.clear()
    await message.answer(f"✅ Vakansiya yaratildi: <b>{v.title}</b>", parse_mode="HTML")


@router.callback_query(lambda c: c.data.startswith("admin_vacancy_toggle:"))
async def toggle_vac(callback: CallbackQuery):
    vacancy_id = int(callback.data.split(":")[1])
    v = await get_vacancy(vacancy_id)
    await toggle_vacancy(vacancy_id, not v.active)
    status = "ochildi" if not v.active else "yopildi"
    await callback.message.answer(
        f"✅ Vakansiya {status}: <b>{v.title}</b>", parse_mode="HTML"
    )
    await callback.answer()


@router.callback_query(lambda c: c.data.startswith("admin_vacancy_delete:"))
async def vacancy_delete_ask(callback: CallbackQuery):
    v = await get_vacancy(int(callback.data.split(":")[1]))
    await callback.message.answer(
        f"⚠️ <b>{v.title}</b> vakansiyasini o'chirishni tasdiqlaysizmi?\n"
        f"<i>Bog'liq barcha arizalar ham o'chadi!</i>",
        parse_mode="HTML",
        reply_markup=vacancy_delete_confirm_keyboard(v.id)
    )
    await callback.answer()


@router.callback_query(lambda c: c.data.startswith("admin_vacancy_confirm_del:"))
async def vacancy_delete_confirm(callback: CallbackQuery):
    vacancy_id = int(callback.data.split(":")[1])
    v = await get_vacancy(vacancy_id)
    title = v.title if v else vacancy_id
    await delete_vacancy(vacancy_id)
    await callback.message.answer(f"🗑 <b>{title}</b> o'chirildi.", parse_mode="HTML")
    await callback.answer()


# ── Vakansiyani guruhga yuborish ──────────────────────────────────────────

def _vacancy_post_text(v) -> str:
    return (
        f"💼 <b>{v.title}</b>\n\n"
        f"📋 <b>Talablar:</b>\n{v.requirements or '—'}\n\n"
        f"💰 <b>Ish haqi:</b> {v.salary or 'Kelishiladi'}"
    )


@router.callback_query(lambda c: c.data == "vac_post:menu")
async def vacancy_post_menu(callback: CallbackQuery):
    role = await get_role(callback.from_user.id)
    if not is_hr(role):
        await callback.answer("❌ Ruxsat yo'q.")
        return
    vacancies = await get_all_vacancies()
    active = [v for v in vacancies if v.active]
    if not active:
        await callback.answer("Faol vakansiya yo'q.", show_alert=True)
        return
    await callback.message.answer(
        "📢 <b>Vakansiyalarni guruhga yuborish</b>\n\n"
        "Barchasini yoki bittasini tanlang:",
        parse_mode="HTML",
        reply_markup=vacancy_post_menu_keyboard(vacancies)
    )
    await callback.answer()


@router.callback_query(lambda c: c.data == "vac_post:all")
async def vacancy_post_all(callback: CallbackQuery, bot: Bot):
    role = await get_role(callback.from_user.id)
    if not is_hr(role):
        await callback.answer("❌ Ruxsat yo'q.")
        return
    group_id_str = await get_setting("apps_group_id")
    if not group_id_str:
        await callback.answer(
            "❌ Avval Sozlamalar → Arizalar guruhi ni o'rnating.",
            show_alert=True
        )
        return
    try:
        group_id = int(group_id_str)
    except ValueError:
        await callback.answer("❌ Guruh ID xato.", show_alert=True)
        return

    vacancies = [v for v in await get_all_vacancies() if v.active]
    if not vacancies:
        await callback.answer("Faol vakansiya yo'q.", show_alert=True)
        return

    await callback.answer("Yuborilmoqda…")
    sent = 0
    for v in vacancies:
        try:
            await bot.send_message(group_id, _vacancy_post_text(v), parse_mode="HTML")
            sent += 1
        except Exception:
            pass
    await callback.message.answer(
        f"✅ {sent} ta vakansiya guruhga yuborildi."
    )


@router.callback_query(lambda c: c.data.startswith("vac_post:one:"))
async def vacancy_post_one(callback: CallbackQuery, bot: Bot):
    role = await get_role(callback.from_user.id)
    if not is_hr(role):
        await callback.answer("❌ Ruxsat yo'q.")
        return
    group_id_str = await get_setting("apps_group_id")
    if not group_id_str:
        await callback.answer(
            "❌ Avval Sozlamalar → Arizalar guruhi ni o'rnating.",
            show_alert=True
        )
        return
    try:
        group_id = int(group_id_str)
    except ValueError:
        await callback.answer("❌ Guruh ID xato.", show_alert=True)
        return

    vacancy_id = int(callback.data.split(":")[2])
    v = await get_vacancy(vacancy_id)
    if not v:
        await callback.answer("Vakansiya topilmadi.", show_alert=True)
        return
    try:
        await bot.send_message(group_id, _vacancy_post_text(v), parse_mode="HTML")
        await callback.message.answer(f"✅ <b>{v.title}</b> guruhga yuborildi.", parse_mode="HTML")
    except Exception as e:
        await callback.message.answer(f"❌ Yuborilmadi: <code>{e}</code>", parse_mode="HTML")
    await callback.answer()


# ── Vakansiya tahrirlash ───────────────────────────────────────────────────

FIELD_LABELS = {
    "title":        "📝 Nomi",
    "requirements": "📋 Talablar",
    "salary":       "💰 Ish haqi",
}


@router.callback_query(lambda c: c.data.startswith("admin_vacancy_edit:"))
async def vacancy_edit_start(callback: CallbackQuery, state: FSMContext):
    role = await get_role(callback.from_user.id)
    if not is_hr(role):
        await callback.answer("❌ Ruxsat yo'q.")
        return
    vacancy_id = int(callback.data.split(":")[1])
    v = await get_vacancy(vacancy_id)
    await state.update_data(edit_vacancy_id=vacancy_id)
    await state.set_state(EditVacancyState.field)
    await callback.message.answer(
        f"✏️ <b>{v.title}</b> — qaysi maydonni tahrirlaysiz?",
        parse_mode="HTML",
        reply_markup=vacancy_edit_field_keyboard(vacancy_id)
    )
    await callback.answer()


@router.callback_query(EditVacancyState.field, lambda c: c.data.startswith("vedit_field:"))
async def vacancy_edit_field(callback: CallbackQuery, state: FSMContext):
    _, vacancy_id, field = callback.data.split(":")
    await state.update_data(edit_field=field)
    await state.set_state(EditVacancyState.value)
    await callback.message.answer(
        f"{FIELD_LABELS.get(field, field)} uchun yangi matn kiriting:"
    )
    await callback.answer()


@router.message(EditVacancyState.value)
async def vacancy_edit_save(message: Message, state: FSMContext):
    data = await state.get_data()
    vacancy_id = data["edit_vacancy_id"]
    field = data["edit_field"]
    await state.clear()
    v = await update_vacancy(vacancy_id, **{field: message.text.strip()})
    await message.answer(
        f"✅ <b>{v.title}</b> yangilandi!\n"
        f"{FIELD_LABELS.get(field, field)}: {message.text.strip()}",
        parse_mode="HTML",
        reply_markup=admin_vacancy_detail_keyboard(v.id, v.active)
    )


# ── Arizalar ───────────────────────────────────────────────────────────────

@router.callback_query(lambda c: c.data == "admin:applications")
async def admin_applications(callback: CallbackQuery):
    role = await get_role(callback.from_user.id)
    if not is_hr(role):
        await callback.answer("❌ Ruxsat yo'q.")
        return
    vacancies = await get_all_vacancies()
    from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
    buttons = [
        [InlineKeyboardButton(text=v.title, callback_data=f"admin_apps:{v.id}")]
        for v in vacancies
    ]
    buttons.append([InlineKeyboardButton(text="📁 Barcha arizalar", callback_data="admin_apps:all")])
    buttons.append([InlineKeyboardButton(text="🔍 Tartib bo'yicha qidirish", callback_data="admin_app_search")])
    buttons.append([InlineKeyboardButton(text="📢 Guruhga yuborish",          callback_data="app_post:menu")])
    await callback.message.answer(
        "Vakansiya tanlang:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons)
    )
    await callback.answer()


APPS_PAGE_SIZE = 10


def _app_card_text(app, tartib: int, vacancy) -> str:
    yosh = app.age or app.birth_year or '—'
    return (
        f"📁 <b>Ariza #{app.id}</b> | 🔢 Tartib: <b>{tartib}</b>\n"
        f"👤 {app.full_name}\n"
        f"📱 {app.phone}\n"
        f"🎂 Yosh: {yosh}\n"
        f"📍 Qayerdan: {app.address or '—'}\n"
        f"🗣 Tillar: {app.languages or '—'}\n"
        f"💼 Lavozim: {vacancy.title if vacancy else '—'}\n"
        f"🏢 Ish tajribasi: {app.experience or '—'}\n"
        f"🎓 Ma'lumot: {app.education or '—'}\n"
        f"✨ Qo'shimcha ko'nikmalar: {app.additional_skills or '—'}"
    )


def _app_card_keyboard(app):
    from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
    rows = []
    media_row = []
    if app.photo_file_id:
        media_row.append(InlineKeyboardButton(text="📷 Rasm", callback_data=f"get_photo:{app.id}"))
    if app.cv_file_id:
        media_row.append(InlineKeyboardButton(text="📄 CV", callback_data=f"get_cv:{app.id}"))
    if media_row:
        rows.append(media_row)
    rows.append([InlineKeyboardButton(text="🗑 O'chirish", callback_data=f"app_del:{app.id}")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


@router.callback_query(lambda c: c.data.startswith("admin_apps:"))
async def show_applications(callback: CallbackQuery, bot: Bot):
    parts = callback.data.split(":")
    val = parts[1]
    page = int(parts[2]) if len(parts) > 2 else 0

    vid = None if val == "all" else int(val)
    apps = await get_applications(vid)
    if not apps:
        await callback.message.answer("Arizalar yo'q.")
        await callback.answer()
        return

    total = len(apps)
    total_pages = (total + APPS_PAGE_SIZE - 1) // APPS_PAGE_SIZE
    page = max(0, min(page, total_pages - 1))
    start = page * APPS_PAGE_SIZE
    end = start + APPS_PAGE_SIZE
    page_apps = apps[start:end]

    from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton

    for idx_in_page, app in enumerate(page_apps):
        tartib = total - (start + idx_in_page)
        v = await get_vacancy(app.vacancy_id) if app.vacancy_id else None
        await callback.message.answer(
            _app_card_text(app, tartib, v),
            parse_mode="HTML",
            reply_markup=_app_card_keyboard(app)
        )

    nav = []
    if page > 0:
        nav.append(InlineKeyboardButton(text="◀️ Oldingi", callback_data=f"admin_apps:{val}:{page - 1}"))
    if end < total:
        nav.append(InlineKeyboardButton(text="Keyingi ▶️", callback_data=f"admin_apps:{val}:{page + 1}"))

    nav_kb = InlineKeyboardMarkup(inline_keyboard=[nav]) if nav else None
    await callback.message.answer(
        f"📄 Sahifa {page + 1}/{total_pages} — jami {total} ta ariza",
        reply_markup=nav_kb
    )
    await callback.answer()


# ── Ariza qidirish (tartib raqami bo'yicha) ───────────────────────────────

@router.callback_query(lambda c: c.data == "admin_app_search")
async def app_search_start(callback: CallbackQuery, state: FSMContext):
    role = await get_role(callback.from_user.id)
    if not is_hr(role):
        await callback.answer("❌ Ruxsat yo'q.")
        return
    apps = await get_applications()
    if not apps:
        await callback.answer("Arizalar yo'q.", show_alert=True)
        return
    await state.set_state(SearchApplicationState.tartib)
    await callback.message.answer(
        f"🔍 Qidirilayotgan <b>tartib raqamini</b> (1 dan {len(apps)} gacha) "
        f"yoki <b>ism familiyani</b> kiriting:",
        parse_mode="HTML"
    )
    await callback.answer()


@router.message(SearchApplicationState.tartib)
async def app_search_run(message: Message, state: FSMContext):
    text = message.text.strip()
    apps = await get_applications()
    total = len(apps)
    await state.clear()

    if text.isdigit():
        tartib = int(text)
        if not (1 <= tartib <= total):
            await message.answer(f"⚠️ Tartib {1} dan {total} gacha bo'lishi kerak.")
            return
        app = apps[total - tartib]
        v = await get_vacancy(app.vacancy_id) if app.vacancy_id else None
        await message.answer(
            _app_card_text(app, tartib, v),
            parse_mode="HTML",
            reply_markup=_app_card_keyboard(app)
        )
        return

    query = text.lower()
    matches = [
        (idx, a) for idx, a in enumerate(apps)
        if a.full_name and query in a.full_name.lower()
    ]
    if not matches:
        await message.answer(f"❌ \"{text}\" bo'yicha ariza topilmadi.")
        return
    await message.answer(f"🔍 <b>{len(matches)} ta natija topildi:</b>", parse_mode="HTML")
    for idx, app in matches[:20]:
        tartib = total - idx
        v = await get_vacancy(app.vacancy_id) if app.vacancy_id else None
        await message.answer(
            _app_card_text(app, tartib, v),
            parse_mode="HTML",
            reply_markup=_app_card_keyboard(app)
        )
    if len(matches) > 20:
        await message.answer(f"… va yana {len(matches) - 20} ta natija. Aniqroq qidiring.")


# ── Ariza o'chirish ────────────────────────────────────────────────────────

@router.callback_query(lambda c: c.data.startswith("app_del:"))
async def app_delete_ask(callback: CallbackQuery):
    role = await get_role(callback.from_user.id)
    if not is_hr(role):
        await callback.answer("❌ Ruxsat yo'q.")
        return
    app_id = int(callback.data.split(":")[1])
    app = await get_application(app_id)
    if not app:
        await callback.answer("Ariza topilmadi.", show_alert=True)
        return
    from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
    kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="✅ Ha, o'chirish", callback_data=f"app_del_yes:{app_id}"),
        InlineKeyboardButton(text="❌ Bekor", callback_data=f"app_del_no:{app_id}"),
    ]])
    await callback.message.answer(
        f"🗑 <b>Ariza #{app_id}</b> ({app.full_name}) ni o'chirishni tasdiqlaysizmi?",
        parse_mode="HTML",
        reply_markup=kb
    )
    await callback.answer()


@router.callback_query(lambda c: c.data.startswith("app_del_yes:"))
async def app_delete_confirm(callback: CallbackQuery):
    role = await get_role(callback.from_user.id)
    if not is_hr(role):
        await callback.answer("❌ Ruxsat yo'q.")
        return
    app_id = int(callback.data.split(":")[1])
    ok = await delete_application(app_id)
    if ok:
        await callback.message.edit_text(f"🗑 Ariza #{app_id} o'chirildi.")
    else:
        await callback.message.edit_text("⚠️ Ariza topilmadi.")
    await callback.answer()


@router.callback_query(lambda c: c.data.startswith("app_del_no:"))
async def app_delete_cancel(callback: CallbackQuery):
    await callback.message.edit_text("❌ O'chirish bekor qilindi.")
    await callback.answer()


# ── Arizalarni guruhga yuborish ────────────────────────────────────────────

def _application_post_text(app, tartib: int, vacancy) -> str:
    from datetime import datetime
    yosh = app.age or app.birth_year or '—'
    created = '—'
    if app.created_at:
        if isinstance(app.created_at, datetime):
            created = app.created_at.strftime("%Y-%m-%d %H:%M")
        else:
            created = str(app.created_at)
    return (
        f"📁 <b>Ariza #{app.id}</b> | 🔢 Tartib: <b>{tartib}</b>\n"
        f"🕐 Topshirilgan: {created}\n"
        f"👤 {app.full_name or '—'}\n"
        f"📱 {app.phone or '—'}\n"
        f"🎂 Yosh: {yosh}\n"
        f"📍 Qayerdan: {app.address or '—'}\n"
        f"🗣 Tillar: {app.languages or '—'}\n"
        f"💼 Lavozim: {vacancy.title if vacancy else '—'}\n"
        f"🏢 Ish tajribasi: {app.experience or '—'}\n"
        f"🎓 Ma'lumot: {app.education or '—'}\n"
        f"✨ Qo'shimcha ko'nikmalar: {app.additional_skills or '—'}"
    )


@router.callback_query(lambda c: c.data == "app_post:menu")
async def app_post_menu(callback: CallbackQuery):
    role = await get_role(callback.from_user.id)
    if not is_hr(role):
        await callback.answer("❌ Ruxsat yo'q.")
        return
    vacancies = await get_all_vacancies()
    await callback.message.answer(
        "📢 <b>Arizalarni guruhga yuborish</b>\n\n"
        "Hammasini yoki ma'lum vakansiya bo'yicha tanlang:",
        parse_mode="HTML",
        reply_markup=applications_post_menu_keyboard(vacancies)
    )
    await callback.answer()


async def _send_apps_to_group(callback: CallbackQuery, bot: Bot, vacancy_id: int | None):
    role = await get_role(callback.from_user.id)
    if not is_hr(role):
        await callback.answer("❌ Ruxsat yo'q.")
        return
    group_id_str = await get_setting("apps_group_id")
    if not group_id_str:
        await callback.answer(
            "❌ Avval Sozlamalar → Arizalar guruhi ni o'rnating.",
            show_alert=True
        )
        return
    try:
        group_id = int(group_id_str)
    except ValueError:
        await callback.answer("❌ Guruh ID xato.", show_alert=True)
        return

    apps = await get_applications(vacancy_id)
    if not apps:
        await callback.answer("Ariza yo'q.", show_alert=True)
        return

    await callback.answer(f"Yuborilmoqda… ({len(apps)} ta)")
    total = len(apps)
    sent = 0
    failed = 0
    import asyncio
    from aiogram.exceptions import TelegramRetryAfter, TelegramAPIError

    async def _send_one(app, text):
        if app.photo_file_id:
            await bot.send_photo(group_id, photo=app.photo_file_id,
                                 caption=text, parse_mode="HTML")
        else:
            await bot.send_message(group_id, text, parse_mode="HTML")

    for idx, app in enumerate(apps):
        tartib = total - idx
        v = await get_vacancy(app.vacancy_id) if app.vacancy_id else None
        text = _application_post_text(app, tartib, v)

        ok = False
        for attempt in range(3):
            try:
                await _send_one(app, text)
                ok = True
                break
            except TelegramRetryAfter as e:
                # Telegram aytgan vaqt + 1 sek qo'shimcha kutamiz, keyin qayta urinamiz
                await asyncio.sleep(e.retry_after + 1)
            except TelegramAPIError:
                break  # boshqa API xato — qayta urinmaymiz
            except Exception:
                break

        if ok:
            sent += 1
        else:
            failed += 1

    summary = f"✅ {sent}/{total} ta ariza guruhga yuborildi."
    if failed:
        summary += f"\n⚠️ {failed} tasi yuborilmadi (xatolik)."
    await callback.message.answer(summary)


@router.callback_query(lambda c: c.data == "app_post:all")
async def app_post_all(callback: CallbackQuery, bot: Bot):
    await _send_apps_to_group(callback, bot, vacancy_id=None)


@router.callback_query(lambda c: c.data.startswith("app_post:vac:"))
async def app_post_one_vacancy(callback: CallbackQuery, bot: Bot):
    vid = int(callback.data.split(":")[2])
    await _send_apps_to_group(callback, bot, vacancy_id=vid)


@router.callback_query(lambda c: c.data.startswith("get_cv:"))
async def get_cv(callback: CallbackQuery, bot: Bot):
    app_id = int(callback.data.split(":")[1])
    apps = await get_applications()
    app = next((a for a in apps if a.id == app_id), None)
    if not app or not app.cv_file_id:
        await callback.answer("CV topilmadi.")
        return
    await bot.send_document(
        callback.from_user.id,
        document=app.cv_file_id,
        caption=f"📄 CV — {app.full_name}"
    )
    await callback.answer()


@router.callback_query(lambda c: c.data.startswith("get_photo:"))
async def get_photo(callback: CallbackQuery, bot: Bot):
    app_id = int(callback.data.split(":")[1])
    apps = await get_applications()
    app = next((a for a in apps if a.id == app_id), None)
    if not app or not app.photo_file_id:
        await callback.answer("Rasm topilmadi.")
        return
    await bot.send_photo(
        callback.from_user.id,
        photo=app.photo_file_id,
        caption=f"📷 {app.full_name}"
    )
    await callback.answer()


# ── Excel eksport ──────────────────────────────────────────────────────────

EXCEL_HEADERS = [
    "Tartib", "Ariza №", "Topshirilgan vaqt", "Ism familiya", "Telefon", "Yosh",
    "Qayerdan", "Tillar", "Lavozim", "Ish tajribasi", "Ma'lumot",
    "Qo'shimcha ko'nikmalar", "Rasm", "CV",
]

_BAD_SHEET_CHARS = set('\\/?*[]:')


def _safe_sheet_name(name: str, used: set[str]) -> str:
    clean = "".join("_" if c in _BAD_SHEET_CHARS else c for c in (name or "Belgilanmagan"))
    clean = clean.strip() or "Belgilanmagan"
    clean = clean[:31]
    base = clean
    i = 2
    while clean in used:
        suffix = f" ({i})"
        clean = (base[: 31 - len(suffix)]) + suffix
        i += 1
    used.add(clean)
    return clean


@router.callback_query(lambda c: c.data == "admin:export")
async def export_applications_xlsx(callback: CallbackQuery, bot: Bot):
    role = await get_role(callback.from_user.id)
    if not is_hr(role):
        await callback.answer("❌ Ruxsat yo'q.")
        return

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from aiogram.types import BufferedInputFile
    from io import BytesIO
    from datetime import datetime

    apps = await get_applications()
    if not apps:
        await callback.answer("Arizalar yo'q.", show_alert=True)
        return

    await callback.answer("Tayyorlanmoqda…")

    vacancies = await get_all_vacancies()
    vacancy_map = {v.id: v for v in vacancies}

    grouped: dict[int | None, list] = {}
    for a in apps:
        grouped.setdefault(a.vacancy_id, []).append(a)

    wb = Workbook()
    wb.remove(wb.active)
    used_names: set[str] = set()

    header_font = Font(bold=True, color="000000", size=12)
    header_fill = PatternFill(start_color="FFE599", end_color="FFE599", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_side   = Side(border_style="thin", color="000000")
    header_border = Border(top=thin_side, bottom=thin_side, left=thin_side, right=thin_side)

    ordered_vids = [v.id for v in vacancies if v.id in grouped] + [vid for vid in grouped if vid not in vacancy_map]

    for vid in ordered_vids:
        rows = sorted(
            grouped[vid],
            key=lambda a: a.created_at or datetime.min
        )
        v = vacancy_map.get(vid) if vid else None
        sheet_name = _safe_sheet_name(v.title if v else "Belgilanmagan", used_names)
        ws = wb.create_sheet(title=sheet_name)

        ws.append(EXCEL_HEADERS)
        for col_idx, _ in enumerate(EXCEL_HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = header_border

        for tartib, a in enumerate(rows, start=1):
            yosh = a.age or a.birth_year or ""
            created = ""
            if a.created_at:
                created = a.created_at.strftime("%Y-%m-%d %H:%M") if isinstance(a.created_at, datetime) else str(a.created_at)
            ws.append([
                tartib,
                a.id,
                created,
                a.full_name or "",
                a.phone or "",
                yosh,
                a.address or "",
                a.languages or "",
                v.title if v else "",
                a.experience or "",
                a.education or "",
                a.additional_skills or "",
                "Bor" if a.photo_file_id else "—",
                "Bor" if a.cv_file_id else "—",
            ])

        widths = [8, 10, 18, 26, 16, 6, 22, 22, 22, 30, 18, 30, 8, 8]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"

    if not wb.sheetnames:
        wb.create_sheet("Arizalar")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"arizalar_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    await bot.send_document(
        callback.from_user.id,
        document=BufferedInputFile(buf.read(), filename=filename),
        caption=f"📥 Jami {len(apps)} ta ariza, {len(ordered_vids)} ta vakansiya bo'yicha."
    )


# ── Adminlar boshqaruvi ────────────────────────────────────────────────────

@router.message(Command("addadmin"))
async def add_admin_cmd(message: Message):
    if message.from_user.id != SUPER_ADMIN_ID:
        await message.answer("❌ Faqat super admin uchun.")
        return
    parts = message.text.split()
    if len(parts) < 3:
        await message.answer("Foydalanish: /addadmin <telegram_id> <role>\nRollar: hr_admin, super_admin")
        return
    try:
        tid, role = int(parts[1]), parts[2]
    except ValueError:
        await message.answer("Noto'g'ri telegram_id.")
        return
    await add_admin(tid, full_name="—", role=role, added_by=message.from_user.id)
    await message.answer(f"✅ Admin qo'shildi: {tid} | Rol: {role}")


@router.message(Command("removeadmin"))
async def remove_admin_cmd(message: Message):
    if message.from_user.id != SUPER_ADMIN_ID:
        await message.answer("❌ Faqat super admin uchun.")
        return
    parts = message.text.split()
    if len(parts) < 2:
        await message.answer("Foydalanish: /removeadmin <telegram_id>")
        return
    await remove_admin(int(parts[1]))
    await message.answer(f"✅ Admin o'chirildi: {parts[1]}")


@router.callback_query(lambda c: c.data == "admin:admins")
async def list_admins(callback: CallbackQuery):
    if callback.from_user.id != SUPER_ADMIN_ID:
        await callback.answer("❌ Ruxsat yo'q.")
        return
    admins = await get_all_admins()
    await callback.message.answer(
        f"👥 <b>Adminlar ro'yxati</b> ({len(admins)} ta):",
        parse_mode="HTML",
        reply_markup=admin_list_keyboard(admins)
    )
    await callback.answer()


@router.callback_query(lambda c: c.data.startswith("admin_detail:"))
async def admin_detail(callback: CallbackQuery):
    if callback.from_user.id != SUPER_ADMIN_ID:
        await callback.answer("❌ Ruxsat yo'q.")
        return
    tid = int(callback.data.split(":")[1])
    admin = await get_admin(tid)
    if not admin:
        await callback.answer("Admin topilmadi.")
        return
    text = (
        f"👤 <b>Admin ma'lumotlari</b>\n\n"
        f"🆔 ID: <code>{admin.telegram_id}</code>\n"
        f"📛 Ism: {admin.full_name or '—'}\n"
        f"🎭 Rol: {ROLE_LABELS.get(admin.role, admin.role)}\n"
        f"📅 Qo'shilgan: {str(admin.created_at)[:10] if admin.created_at else '—'}"
    )
    await callback.message.answer(
        text, parse_mode="HTML",
        reply_markup=admin_detail_keyboard(tid)
    )
    await callback.answer()


@router.callback_query(lambda c: c.data == "admin_add:start")
async def admin_add_start(callback: CallbackQuery, state: FSMContext):
    if callback.from_user.id != SUPER_ADMIN_ID:
        await callback.answer("❌ Ruxsat yo'q.")
        return
    await state.set_state(AddAdminState.telegram_id)
    await callback.message.answer(
        "➕ <b>Yangi admin qo'shish</b>\n\n"
        "Admin Telegram ID sini kiriting:\n"
        "<i>(@userinfobot orqali bilib olish mumkin)</i>",
        parse_mode="HTML"
    )
    await callback.answer()


@router.message(AddAdminState.telegram_id)
async def admin_add_get_id(message: Message, state: FSMContext):
    if not message.text.strip().lstrip("-").isdigit():
        await message.answer("❌ Raqam kiriting (masalan: 123456789):")
        return
    await state.update_data(new_admin_id=int(message.text.strip()))
    await state.set_state(AddAdminState.role)
    await message.answer("Rolni tanlang:", reply_markup=admin_roles_keyboard("new_admin_role"))


@router.callback_query(AddAdminState.role, lambda c: c.data.startswith("new_admin_role:"))
async def admin_add_set_role(callback: CallbackQuery, state: FSMContext, bot: Bot):
    role = callback.data.split(":")[1]
    data = await state.get_data()
    tid  = data["new_admin_id"]
    await state.clear()
    full_name = "—"
    try:
        chat = await bot.get_chat(tid)
        full_name = chat.full_name or "—"
    except Exception:
        pass
    await add_admin(tid, full_name=full_name, role=role, added_by=callback.from_user.id)
    await callback.message.answer(
        f"✅ Admin qo'shildi!\n"
        f"🆔 <code>{tid}</code> | {full_name} | {ROLE_LABELS.get(role, role)}",
        parse_mode="HTML"
    )
    await callback.answer()


@router.callback_query(lambda c: c.data.startswith("admin_edit_role:"))
async def admin_edit_role_start(callback: CallbackQuery, state: FSMContext):
    if callback.from_user.id != SUPER_ADMIN_ID:
        await callback.answer("❌ Ruxsat yo'q.")
        return
    tid = int(callback.data.split(":")[1])
    await state.update_data(edit_admin_id=tid)
    await state.set_state(EditAdminState.role)
    await callback.message.answer(
        f"Yangi rolni tanlang (ID: <code>{tid}</code>):",
        parse_mode="HTML",
        reply_markup=admin_roles_keyboard("set_admin_role")
    )
    await callback.answer()


@router.callback_query(EditAdminState.role, lambda c: c.data.startswith("set_admin_role:"))
async def admin_edit_role_confirm(callback: CallbackQuery, state: FSMContext):
    role = callback.data.split(":")[1]
    data = await state.get_data()
    await state.clear()
    await update_admin_role(data["edit_admin_id"], role)
    await callback.message.answer(
        f"✅ Rol yangilandi: {ROLE_LABELS.get(role, role)}", parse_mode="HTML"
    )
    await callback.answer()


@router.callback_query(lambda c: c.data.startswith("admin_remove:"))
async def admin_remove_ask(callback: CallbackQuery):
    if callback.from_user.id != SUPER_ADMIN_ID:
        await callback.answer("❌ Ruxsat yo'q.")
        return
    tid   = int(callback.data.split(":")[1])
    admin = await get_admin(tid)
    name  = admin.full_name if admin else tid
    await callback.message.answer(
        f"⚠️ <b>{name}</b> adminni o'chirishni tasdiqlaysizmi?",
        parse_mode="HTML",
        reply_markup=admin_remove_confirm_keyboard(tid)
    )
    await callback.answer()


@router.callback_query(lambda c: c.data.startswith("admin_remove_confirm:"))
async def admin_remove_confirm(callback: CallbackQuery):
    if callback.from_user.id != SUPER_ADMIN_ID:
        await callback.answer("❌ Ruxsat yo'q.")
        return
    tid = int(callback.data.split(":")[1])
    await remove_admin(tid)
    await callback.message.answer(f"✅ Admin <code>{tid}</code> o'chirildi.", parse_mode="HTML")
    await callback.answer()


# ── Bot sozlamalari ────────────────────────────────────────────────────────

async def _settings_text() -> str:
    channel_link = await get_setting("channel_link")
    instagram    = await get_setting("instagram_link")
    group_title  = await get_setting("apps_group_title")
    group_id     = await get_setting("apps_group_id")
    not_set      = "— (o'rnatilmagan)"
    group_display = f"{group_title} (<code>{group_id}</code>)" if group_id else not_set
    return (
        "⚙️ <b>Bot sozlamalari</b>\n\n"
        f"📡 Kanal: {channel_link or not_set}\n"
        f"📸 Instagram: {instagram or not_set}\n"
        f"📥 Arizalar guruhi: {group_display}\n\n"
        "<i>Kanal o'rnatilsa — foydalanuvchilar botdan foydalanishdan oldin "
        "kanalga a'zo bo'lishi shart bo'ladi.\n"
        "Arizalar guruhi o'rnatilsa — har bir yangi ariza shu guruhga ham yuboriladi.</i>"
    )


@router.callback_query(lambda c: c.data == "admin:settings")
async def admin_settings(callback: CallbackQuery):
    if callback.from_user.id != SUPER_ADMIN_ID:
        await callback.answer("❌ Ruxsat yo'q.")
        return
    await callback.message.edit_text(await _settings_text(), parse_mode="HTML",
                                     reply_markup=admin_settings_keyboard())
    await callback.answer()


@router.callback_query(lambda c: c.data == "settings:channel")
async def settings_channel_start(callback: CallbackQuery, state: FSMContext):
    if callback.from_user.id != SUPER_ADMIN_ID:
        await callback.answer("❌ Ruxsat yo'q.")
        return
    await state.set_state(BotSettingsState.channel)
    await callback.message.answer(
        "📡 <b>Kanal o'rnatish</b>\n\n"
        "Kanal username yoki linkini yuboring:\n"
        "<code>@NurlidiyorResidence</code>\n"
        "<code>https://t.me/NurlidiyorResidence</code>\n\n"
        "<i>Eslatma: bot kanalda admin bo'lishi kerak!</i>",
        parse_mode="HTML"
    )
    await callback.answer()


@router.message(BotSettingsState.channel)
async def settings_channel_save(message: Message, state: FSMContext, bot: Bot):
    raw = message.text.strip()

    # Username yoki linkdan username ajratib olamiz
    username = raw
    if raw.startswith("https://t.me/"):
        username = "@" + raw.split("t.me/")[-1].strip("/")
    elif not raw.startswith("@"):
        username = "@" + raw

    try:
        chat = await bot.get_chat(username)
        chat_id   = chat.id
        chat_link = f"https://t.me/{chat.username}" if chat.username else raw
    except Exception as e:
        await message.answer(
            f"❌ Kanal topilmadi yoki bot kanalga qo'shilmagan.\n"
            f"Xatolik: <code>{e}</code>\n\n"
            "Iltimos, botni kanalga <b>admin</b> qiling va qayta urinib ko'ring.",
            parse_mode="HTML"
        )
        return

    await set_setting("channel_id",   str(chat_id))
    await set_setting("channel_link", chat_link)
    await state.clear()
    await message.answer(
        f"✅ Kanal o'rnatildi!\n"
        f"📡 <b>{chat.title}</b>\n"
        f"🔗 {chat_link}",
        parse_mode="HTML",
        reply_markup=admin_settings_keyboard()
    )


@router.callback_query(lambda c: c.data == "settings:instagram")
async def settings_instagram_start(callback: CallbackQuery, state: FSMContext):
    if callback.from_user.id != SUPER_ADMIN_ID:
        await callback.answer("❌ Ruxsat yo'q.")
        return
    await state.set_state(BotSettingsState.instagram)
    await callback.message.answer(
        "📸 <b>Instagram linkini o'rnatish</b>\n\n"
        "Instagram sahifangiz to'liq linkini yuboring:\n"
        "<code>https://instagram.com/nuriddinbuildings</code>",
        parse_mode="HTML"
    )
    await callback.answer()


@router.message(BotSettingsState.instagram)
async def settings_instagram_save(message: Message, state: FSMContext):
    link = message.text.strip()
    await set_setting("instagram_link", link)
    await state.clear()
    await message.answer(
        f"✅ Instagram linki saqlandi!\n📸 {link}",
        reply_markup=admin_settings_keyboard()
    )


@router.callback_query(lambda c: c.data == "settings:clear_channel")
async def settings_clear_channel(callback: CallbackQuery):
    if callback.from_user.id != SUPER_ADMIN_ID:
        await callback.answer("❌ Ruxsat yo'q.")
        return
    await set_setting("channel_id",   None)
    await set_setting("channel_link", None)
    await callback.answer("✅ Kanal sozlamasi o'chirildi.", show_alert=True)
    await callback.message.edit_text(
        await _settings_text(), parse_mode="HTML",
        reply_markup=admin_settings_keyboard()
    )


@router.callback_query(lambda c: c.data == "settings:apps_group")
async def settings_apps_group_help(callback: CallbackQuery):
    if callback.from_user.id != SUPER_ADMIN_ID:
        await callback.answer("❌ Ruxsat yo'q.")
        return
    await callback.message.answer(
        "📥 <b>Arizalar guruhini sozlash</b>\n\n"
        "1. Botni guruhga qo'shing (super_admin sifatida)\n"
        "2. Guruh ichida <code>/set_apps_group</code> deb yuboring\n"
        "3. Bot guruh ID'sini avtomatik saqlaydi\n\n"
        "<i>Shundan keyin har bir yangi ariza shu guruhga yuboriladi.</i>",
        parse_mode="HTML"
    )
    await callback.answer()


@router.callback_query(lambda c: c.data == "settings:clear_group")
async def settings_clear_group(callback: CallbackQuery):
    if callback.from_user.id != SUPER_ADMIN_ID:
        await callback.answer("❌ Ruxsat yo'q.")
        return
    await set_setting("apps_group_id",    None)
    await set_setting("apps_group_title", None)
    await callback.answer("✅ Arizalar guruhi o'chirildi.", show_alert=True)
    await callback.message.edit_text(
        await _settings_text(), parse_mode="HTML",
        reply_markup=admin_settings_keyboard()
    )


@router.message(Command("set_apps_group"))
async def set_apps_group_cmd(message: Message):
    if message.from_user.id != SUPER_ADMIN_ID:
        await message.answer("❌ Faqat super admin uchun.")
        return
    if message.chat.type not in ("group", "supergroup"):
        await message.answer("⚠️ Bu buyruq faqat guruh ichida ishlaydi.")
        return
    await set_setting("apps_group_id",    str(message.chat.id))
    await set_setting("apps_group_title", message.chat.title or "Guruh")
    await message.answer(
        f"✅ Arizalar guruhi o'rnatildi!\n"
        f"📥 <b>{message.chat.title}</b>\n"
        f"🆔 <code>{message.chat.id}</code>\n\n"
        f"Endi har bir yangi ariza shu guruhga yuboriladi.",
        parse_mode="HTML"
    )
